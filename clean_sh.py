from openpyxl import load_workbook

def main():
    
    wb2 = load_workbook("komunikat.xlsx")   
    
    sheet2 = wb2['Arkusz1']   
    
    for i in range(1, 300):
        for j in range(2, 100):
                
            sheet2.cell(row = i, column = j).value = None
    
    wb2.save("komunikat.xlsx")
    wb2.close()
    
if __name__ == "__main__":
    main()
