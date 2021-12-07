from PIL import Image, ImageOps
import PIL
import math
from openpyxl import load_workbook

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

def calc_avr(dist_list) -> float:
    
    return sum(dist_list)/len(dist_list)
    

def calc_dev(dist_list, avr):

    suma = 0

    for i in range(len(dist_list)):
        suma += ((dist_list[i] - avr) ** 2)
    
    return math.sqrt(suma/len(dist_list))

def calc_perc_dev(dev, avr):
    return dev/avr * 100

def quick_dist(xl: list, xr: list, n: int, h: int) -> float:
    
    #lewy skrajny
    left_px = xl[-2]

    #prawy skrajny 
    right_px = xr[-2]

    #srodkowy
    middle_px = xl[0]

    diff_list = []

    diff_list.append(abs(left_px[1] - middle_px[1])) # [0] - lm
    diff_list.append(abs(left_px[1] - right_px[1])) # [1] - lr
    diff_list.append(abs(middle_px[1] - right_px[1])) # [2] - mr

    r = round(max(diff_list)/h * 100, 2)
    
    return r

def dystorsja(dist_list):
    
    tollerancy = 1

    a = calc_avr(dist_list)
    d = calc_dev(dist_list, a)
    pd = calc_perc_dev(d, a)

    if(pd < tollerancy):
        #brak dystorsji
        return 2

    else:
        differences = []

        for i in range(1, len(dist_list)):
            differences.append(dist_list[i] - dist_list[i - 1])
        
        s = sum(differences)

        if(s > 0):
            #dystorsja poduszkowa
            return 3
        elif(s < 0):
            #dystorsja beczkowa
            return 1

def processing(n: int):
    
    v_dist_left = []
    v_dist_right = []
    h_dist_up = []
    h_dist_down = []
    
    x_left = []
    x_right = []

    im = Image.open('cropped_images/D' + str(n + 1) + '.png')
    mono = ImageOps.grayscale(im)
    x, y = mono.size

    back_value = 100
    rect_value = 80

    v_half = math.floor(y/2)
    h_half = math.floor(x/2)

    v_factor = -1
    v_pointer = -1
    j = 0

    while(j <= x):
        
        white_passed = False
        h_line = h_half + v_factor * j

        for l in range(v_half):

            px = mono.getpixel((h_line, l))
            if(px < rect_value and not white_passed):
                white_passed = True
            elif(px > back_value and white_passed):
                v_pointer = l
                #obliczanie zagiecia
                if(v_factor == -1):
                    x_left.append((h_line, v_pointer))
                elif(v_factor == 1):
                    x_right.append((h_line, v_pointer))
                break

        if(l >= v_half - 1 and v_factor == -1):
            v_factor = 1
            j = 0
            continue
        if(l >= v_half - 1 and v_factor == 1):
            break

        for i in range(1, (y - v_pointer - 1)):
            z = mono.getpixel((h_line, v_pointer + i))
            if(z < rect_value):
                #print("dist:", i)
                if(v_factor == -1):
                    v_dist_left.append(i)
                elif(v_factor == 1):
                    v_dist_right.append(i)
                break
        j += 1

    h_factor = -1
    h_pointer = -1
    j = 0

    while(j <= y):

        white_passed = False
        v_line = v_half + h_factor * j

        for l in range(h_half):

            px = mono.getpixel((l, v_line))
            if(px < rect_value and not white_passed):
                white_passed = True
            elif(px > back_value and white_passed):
                h_pointer = l
                break

        if(l >= h_half - 1 and h_factor == -1):
            h_factor = 1
            j = 0
            continue
        if(l >= h_half - 1 and h_factor == 1):
            break

        for i in range(1, (x - h_pointer - 1)):
            z = mono.getpixel((h_pointer + i, v_line))
            if(z < rect_value):

                if(h_factor == -1):
                    h_dist_up.append(i)
                elif(h_factor == 1):
                    h_dist_down.append(i)
                break
        
        j += 1

    if(h_dist_down and h_dist_up and v_dist_left and v_dist_right and x_right and x_left):

        out = quick_dist(x_left, x_right, n, y)
        
        h_list = h_dist_up + h_dist_down
        v_list = v_dist_left + v_dist_right
        
        ha = calc_avr(h_list)
        hd = calc_dev(h_list, ha)
        hpd = calc_perc_dev(hd, ha)
        
        va = calc_avr(v_list)
        vd = calc_dev(v_list, va)
        vpd = calc_perc_dev(vd, va)

        output = [hpd, vpd, out]
        
    else:
        output = [None, None, None, None]
           
    return output
        
def main():
    
    
    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 23)
    
    do_sum = 0
    pd = 0
    n = 0
    for i in range(0, 4):
        print("Przetwarzanie: Dystorsje" + str(i + 1) + "...")
        if(processing(i)[0]):
            #hpd
            h = round(processing(i)[0], 3)
            r1 = 123 + i * 4 
            sheet2.cell(row = r1, column = empty_col).value = h
            #vpd
            v = round(processing(i)[1], 3)
            sheet2.cell(row = r1 + 1, column = empty_col).value = v
            #zagięcie procentowe
            z = round(processing(i)[2], 3)
            sheet2.cell(row = r1 + 2, column = empty_col).value = z
            
            message = ""
            
            
            do_sum += z
              
    m = ""
     
    if(-4 < do_sum < 4):
        #brak dystorsji
        m = "Brak dystorsji"
    elif(do_sum < -8):
        #dystorsja poduszkowa
        m = "Dystorsja poduszkowa"
    elif(-8 <= do_sum <= -4): 
        m = "Umiarkowana dystorsja poduszkowa"
    elif(do_sum > 8):
        #dystorsja beczkowa
        m = "Dystorsja beczkowa"
    elif(4 <= do_sum <= 8):
        m = "Umiarkowana dystorsja beczkowa"
    else:
        #błąd
        m = "Błąd"
    
    sheet2.cell(row = 23, column = empty_col).value = m
    
    wb2.save('komunikat.xlsx') 
    print("Dystorsje przetworzone")
    
if __name__ == "__main__":
    main()