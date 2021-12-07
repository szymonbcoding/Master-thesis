from PIL import Image
import PIL
import math

from openpyxl import load_workbook

import glob

def takeThird(elem: tuple) -> float:
    return elem[2]


def openFolder(path):
    for filename in glob.glob(path + '/*.png'):
        img=Image.open(filename)

    return img

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c


def contrast(vmax: int, vmin: int) -> float:
    return ((vmax - vmin)/(vmax + vmin))

def mtf(cf: float, c0: float) -> float:
    return (100 * cf/c0)

def rt(n: int):
    
    print("Przetwarzanie: BW_RT", str(n) + "...")

    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 4)
    
    #maksymalna wysokosc testu rozdzielczosci[mm]
    MAX = 20
    #wysokosc calej planszy testowej [mm]
    h = 366.67

    #zaladowanie całego obrazu planszy
    im1 = openFolder("cropped_photo")

    x_all, y_all = im1.size

    if(0.63 < y_all/x_all < 0.705):
        w = 550
    elif(0.705 <= y_all/x_all < 0.79):
        w = 488.889
    else:
        print("Błędna rozdzielczość planszy.")

    #przeliczenie 20mm (bok testu rozdzielczosci) na piksele
    px_20mm = math.ceil(y_all * MAX/h)

    #zaladowanie obrazu w wersji monochromatycznej
    im2 = Image.open('cropped_images/BW_RT' + str(n) + '.png').convert("L")

    #rozdzielczosc wycinka
    x, y = im2.size
    
    #PETLA TESTU ROZDZIELCZOSCI POZIOMEJ

    #inicjalizacja zmiennych pomocniczych
    hor_res = 0
    v_pointer = -1
    h_pointer = -1

    h_mtf_pointer = -1
    vw_0 = -1
    vb_0 = 256
    
    for i in range(x):
        for j in range(y):
            if(im2.getpixel((i,j)) > vw_0):
                vw_0 = im2.getpixel((i,j))
            elif(im2.getpixel((i,j)) < vb_0):
                vb_0 = im2.getpixel((i,j))

    for i in range(math.floor(y / 2)):
        #resetowanie zmiennych
        hor_edges = 0
        hor_flag = False

        for j in range(x):
              
            if(im2.getpixel((j,i)) > 100 and not hor_flag):

                hor_flag = True
            
            elif(im2.getpixel((j,i)) <= 100 and hor_flag):

                hor_flag = False
                hor_edges += 1
                
                if(h_pointer == -1 and v_pointer != -1):
                    h_pointer = j

            #sprawdzanie czy program naliczyl 5 zbocz 
            if(hor_edges>=5):
                
                if(v_pointer == -1):
                
                    v_pointer = i + px_20mm
                
                hor_res += 1
                h_mtf_pointer = i
                break

    #stosunek poprawnych wierszy (najlepszy mozliwy wynik = 1, najgorszy mozliwy wynik = 0)
    p_row_found = hor_res/px_20mm

    h_mm_resolution = 2 - (1.8 * p_row_found)
    r1 = 31 + n * 11
    sheet2.cell(row = r1, column = empty_col).value = round(h_mm_resolution, 3)
    
    real_h_px_resolution = math.floor(w/h_mm_resolution)

    sheet2.cell(row = r1 + 1, column = empty_col).value = real_h_px_resolution

    # w34
    if(w < 500):
        wmax = 1975
    # w32
    elif(w > 500):
        wmax = 1755
    else:
        wmax = "Blad"

    sheet2.cell(row = r1 + 2, column = empty_col).value = wmax
    
    #MTF
    
    #sprawdzenie o ile spadł kontrast w ostatnim wierszu gdzie rozróżnił
    
    h_vw_f = -1
    h_vb_f = 256

    length_last_row_px = 10 * h_mm_resolution * x_all / w
    
    if(h_pointer + length_last_row_px < x):
        if(not h_mtf_pointer == -1):
            for i in range(h_pointer, math.floor(h_pointer + length_last_row_px)):
                if(im2.getpixel((i, h_mtf_pointer)) > h_vw_f):
                    h_vw_f = im2.getpixel((i, h_mtf_pointer))
                elif(im2.getpixel((i, h_mtf_pointer)) < h_vb_f):
                    h_vb_f = im2.getpixel((i, h_mtf_pointer))

            h_mtf_result = mtf(contrast(h_vw_f, h_vb_f), contrast(vw_0, vb_0))
            
            sheet2.cell(row = r1 + 3, column = empty_col).value = round(h_mtf_result, 2)
        
        else:
            sheet2.cell(row = r1 + 3, column = empty_col).value = "Blad"
    else:
        sheet2.cell(row = r1 + 3, column = empty_col).value = "Blad"
    #PETLA TESTU ROZDZIELCZOSCI PIONOWEJ

    ver_res = 0
    v_pointer = -1
    h_pointer = -1

    
    v_mtf_pointer = -1

    for i in range(x):
        #resetowanie zmiennych
        ver_edges = 0
        ver_flag = False

        for j in range (v_pointer, y):

            if(im2.getpixel((i,j)) <= 100 and not ver_flag):

                ver_flag = True
            
            elif(im2.getpixel((i,j)) > 100 and ver_flag):
                
                ver_flag = False
                ver_edges += 1
            
            if(v_pointer == -1 and h_pointer != -1):
                    
                    v_pointer = j

            #sprawdzanie czy wykryto przynajmniej 5 zbocz
            if(ver_edges>=5):
                
                ver_res += 1
                v_mtf_pointer = j
                h_pointer = i
                break

    p_col_found = ver_res/px_20mm

    v_mm_resolution = 2 - (1.8 * p_col_found)
    sheet2.cell(row = r1 + 5, column = empty_col).value = round(v_mm_resolution, 3)
    
    real_v_px_resolution = math.floor(h/v_mm_resolution)
    sheet2.cell(row = r1 + 6, column = empty_col).value = real_v_px_resolution
    
    # h max
    sheet2.cell(row = r1 + 7, column = empty_col).value = 1316
    #MTF
    
    #sprawdzenie o ile spadł kontrast w ostatnim wierszu gdzie rozróżnił

    v_vw_f = -1
    v_vb_f = 256

    length_last_col_px = 10 * v_mm_resolution * y_all / h
    if(v_pointer + 0.8 *length_last_row_px < y):
        if(not h_pointer == -1):
            for j in range(math.floor(v_mtf_pointer - 0.8 * length_last_col_px), v_mtf_pointer):
                if(im2.getpixel((h_pointer, j)) > v_vw_f):
                    v_vw_f = im2.getpixel((h_pointer, j))
                elif(im2.getpixel((h_pointer, j)) < v_vb_f):
                    v_vb_f = im2.getpixel((h_pointer, j))

            v_mtf_result = mtf(contrast(v_vw_f, v_vb_f), contrast(vw_0, vb_0))

            sheet2.cell(row = r1 + 8, column = empty_col).value = round(v_mtf_result, 2)
        else:
            sheet2.cell(row = r1 + 8, column = empty_col).value = "Blad"
    else:
        sheet2.cell(row = r1 + 8, column = empty_col).value = "Blad"
    
    wb2.save('komunikat.xlsx') 
    print("BW_RT przetworzone")
    
    return (real_h_px_resolution, real_v_px_resolution, real_h_px_resolution * real_v_px_resolution)
    
    
def main(): 

    bw_real_res_list = []
    
    for n in range(1,6):
        bw_real_res_list.append(rt(n))

    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 4)
        
    bw_real_res_list.sort(key = takeThird)
    
    s_v = 0
    s_h = 0
    s_vh = 0
    
    for n in bw_real_res_list:
        s_h += n[0]
        s_v += n[1]
        s_vh += n[2]
    
    #srednia
    
    sheet2.cell(row = 4, column = empty_col).value = round(s_vh/5, 0)
    
    sheet2.cell(row = 5, column = empty_col).value = round(s_h/5, 0)
    
    sheet2.cell(row = 6, column = empty_col).value = round(s_v/5, 0)
    
    #najlepszy wynik
    
    sheet2.cell(row = 7, column = empty_col).value = bw_real_res_list[-1][2]
    
    sheet2.cell(row = 8, column = empty_col).value = bw_real_res_list[-1][0]
    
    sheet2.cell(row = 9, column = empty_col).value = bw_real_res_list[-1][1]
    
    #najgorszy wynik
    
    sheet2.cell(row = 10, column = empty_col).value = bw_real_res_list[0][2]
    
    sheet2.cell(row = 11, column = empty_col).value = bw_real_res_list[0][0]
    
    sheet2.cell(row = 12, column = empty_col).value = bw_real_res_list[0][1]
    
    wb2.save('komunikat.xlsx') 
    print("BW_RT przetworzone")
    
  
if __name__ == "__main__":
    main()
            
