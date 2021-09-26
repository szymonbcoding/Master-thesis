import sys
#import rogi 
from PIL import Image
from PIL.ExifTags import TAGS
from openpyxl import load_workbook
import glob
import os

#własne skrypty 
import test_rozdzielczosci_bw_mtf
import test_rozdzielczosci_gb
import test_rozdzielczosci_rb
import test_rozdzielczosci_rg

import kadrowanie
import wycinanie
import winietowanie
import dystorsje
import ocr
import szum_bw
import szum_rgb
import rozroznialnosc

def cut(fn: str) -> str:
     return fn[6:-4]

def openFolderMI(path: str) -> list:
    
    image_list = []
    
    for filename in glob.glob(path + '/*.JPG'):
        im = Image.open(filename)
        image_list.append(im)
        
    for filename in glob.glob(path + '/*.png'):
        im = Image.open(filename)
        image_list.append(im)

    return image_list

def openFolder(path):
    for filename in glob.glob(path + '/*.png'):
        img=Image.open(filename)

    return img

def rm_files_folder(path):
     
     files = glob.glob(path)
          
     for f in files:
          os.remove(f)

def photo_to_repair(fname):
     
     #zapisuje w pliku tkestowym nazwe zdjecia do powtorzenia
     file1 = open("photos_to_repair.txt", "a")
     
     file1.write(fname + "\n")
     file1.close()

def main():
    #https://www.thepythoncode.com/article/extracting-image-metadata-in-python
    
    im_list = []
    im_list = openFolderMI("photo")
    
    for count, im_bef_crop in enumerate(im_list, start = 1):
          
          print("###")
          print("START PRZETWARZANIA", cut(im_bef_crop.filename))
          print(count, "/", len(im_list))
          print("###")
          
          out = kadrowanie.main(im_bef_crop)
          
          if(out):
          
               wycinanie.main(out)
               
               image = openFolder("cropped_photo")
               
               x, y = image.size
                    
               # extract EXIF data
               exifdata = image.getexif()

               #data = [exifdata[0x0110], exifdata[0xA405], exifdata[0x0100], 
               # exifdata[0x0101], exifdata[0x9202][0]/exifdata[0x9202][1], 
               # exifdata[0x9201][0]/exifdata[0x9201][1], exifdata[0x8827], "-" ]

               wb = load_workbook(filename = 'dane_szczegolowe.xlsx')

               sheet = wb['1_Ustawienia_aparatu']

               empty_row = 0

               for i in range(4, 1000):
                    if(not (sheet.cell(row = i, column = 1).value)):
                         empty_row = i
                         break
                    
               #Model aparatu
               try:
                    model = exifdata[0x0110]
                    sheet.cell(row = empty_row, column = 1).value =  model
               except:
                    model = "-"
                    sheet.cell(row = empty_row, column = 1).value = model
                    
               #Dlugosc ogniskowa
               try:
                    sheet.cell(row = empty_row, column = 2).value = exifdata[0xA405] 
               except:
                    sheet.cell(row = empty_row, column = 2).value = "-"
                    
               #Długość obrazu
               try:
                    sheet.cell(row = empty_row, column = 3).value = exifdata[0x0100] 
               except:
                    sheet.cell(row = empty_row, column = 3).value = x
                    
               #Szerokość obrazu
               try:
                    sheet.cell(row = empty_row, column = 4).value = exifdata[0x0101] 
               except:
                    sheet.cell(row = empty_row, column = 4).value = y
                    
               #Przyslona
               try:
                    sheet.cell(row = empty_row, column = 5).value = exifdata[0x9202][0]/exifdata[0x9202][1] 
               except:
                    sheet.cell(row = empty_row, column = 5).value = "5.6"
                    
               #Czas naswietlania
               try:
                    sheet.cell(row = empty_row, column = 6).value = exifdata[0x9201][0]/exifdata[0x9201][1] 
               except:
                    sheet.cell(row = empty_row, column = 6).value = "1/100"
                    
               #ISO
               try:
                    sheet.cell(row = empty_row, column = 7).value = exifdata[0x8827] 
               except:
                    sheet.cell(row = empty_row, column = 7).value = "200"
                    
               #Model obiektywu
               try:
                    sheet.cell(row = empty_row, column = 8).value = exifdata[0xA434] 
               except:
                    sheet.cell(row = empty_row, column = 8).value = "-"
                    
               #Nazwa zdjecia
               try:
                    sheet.cell(row = empty_row, column = 12).value = cut(im_bef_crop.filename)
               except:
                    sheet.cell(row = empty_row, column = 12).value = "-"
                              
               wb.save('dane_szczegolowe.xlsx')
               wb.close()

               wb2 = load_workbook(filename = 'komunikat.xlsx')

               sheet2 = wb2['Arkusz1']

               empty_col = 0

               for i in range(2, 1000):
                    if(not (sheet2.cell(row = 1, column = i).value)):
                         empty_col = i
                         break
                    
               #model
               sheet2.cell(row = 1, column = empty_col).value = model

               #Nazwa zdjecia
               sheet2.cell(row = 2, column = empty_col).value = cut(im_bef_crop.filename)
               
               #rozdzielczosc deklarowana przez producenta po przycieciu
               sheet2.cell(row = 3, column = empty_col).value = x * y

               wb2.save('komunikat.xlsx')
               wb2.close()
               
               test_rozdzielczosci_bw_mtf.main()
               test_rozdzielczosci_rg.main()
               test_rozdzielczosci_gb.main()
               test_rozdzielczosci_rb.main()
               winietowanie.main()
               dystorsje.main() #do dopracowania
               ocr.main()
               szum_bw.main()
               szum_rgb.main()
               rozroznialnosc.main()
               
               print("###")
               print("KONIEC PRZETWARZANIA", cut(im_bef_crop.filename))
               print("###")
          
          else:
               photo_to_repair(cut(im_bef_crop.filename))
               print("NIEPOWODZENIE:", cut(im_bef_crop.filename) )
               
          # rm_files_folder("crop_pom") 
          # rm_files_folder("cropped_photo")
          # rm_files_folder("cropped_images")
          
     
    print("Koniec całego przetwarzania")
    
if __name__ == "__main__":
    main()