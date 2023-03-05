import sys
#import rogi 
from PIL import Image
from PIL.ExifTags import TAGS
from openpyxl import load_workbook
import glob
import os

#własne skrypty 
import test_rozdzielczosci_bw_mtf
import test_rozdzielczosci_gb
import test_rozdzielczosci_rb
import test_rozdzielczosci_rg

import kadrowanie
import wycinanie
import winietowanie
import dystorsje
import ocr
import szum_bw
import szum_rgb
import rozroznialnosc

def cut(fn: str) -> str:
     return fn[6:-4]

def openFolderMI(path: str) -> list:
    
    image_list = []
    
    for filename in glob.glob(path + '/*.JPG'):
        im = Image.open(filename)
        image_list.append(im)
        
    for filename in glob.glob(path + '/*.png'):
        im = Image.open(filename)
        image_list.append(im)

    return image_list

def openFolder(path):
    for filename in glob.glob(path + '/*.png'):
        img=Image.open(filename)

    return img

def rm_files_folder(path):
     
     files = glob.glob(path)
          
     for f in files:
          os.remove(f)

def photo_to_repair(fname):
     
     #zapisuje w pliku tkestowym nazwe zdjecia do powtorzenia
     file1 = open("photos_to_repair.txt", "a")
     
     file1.write(fname + "\n")
     file1.close()

def main():
    #https://www.thepythoncode.com/article/extracting-image-metadata-in-python
    
    im_list = []
    im_list = openFolderMI("../data/photo")
    
    for count, im_bef_crop in enumerate(im_list, start = 1):
          
          print("###")
          print("START PRZETWARZANIA", cut(im_bef_crop.filename))
          print(count, "/", len(im_list))
          print("###")
          
          out = kadrowanie.main(im_bef_crop)
          
          if(out):
          
               wycinanie.main(out)
               
               image = openFolder("../data/cropped_photo")
               
               x, y = image.size
                    
               # extract EXIF data
               exifdata = image.getexif()

               #data = [exifdata[0x0110], exifdata[0xA405], exifdata[0x0100], 
               # exifdata[0x0101], exifdata[0x9202][0]/exifdata[0x9202][1], 
               # exifdata[0x9201][0]/exifdata[0x9201][1], exifdata[0x8827], "-" ]

               #Model aparatu
               try:
                    model = exifdata[0x0110]
                    
               except:
                    model = "-"

               wb2 = load_workbook(filename = '../data/komunikat.xlsx')

               sheet2 = wb2['Arkusz1']

               empty_col = 0

               for i in range(2, 1000):
                    if(not (sheet2.cell(row = 1, column = i).value)):
                         empty_col = i
                         break
                    
               #model
               sheet2.cell(row = 1, column = empty_col).value = model

               #Nazwa zdjecia
               sheet2.cell(row = 2, column = empty_col).value = cut(im_bef_crop.filename)
               
               #rozdzielczosc deklarowana przez producenta po przycieciu
               sheet2.cell(row = 3, column = empty_col).value = x * y

               wb2.save('../data/komunikat.xlsx')
               wb2.close()
               
               test_rozdzielczosci_bw_mtf.main()
               test_rozdzielczosci_rg.main()
               test_rozdzielczosci_gb.main()
               test_rozdzielczosci_rb.main()
               winietowanie.main()
               dystorsje.main() #do dopracowania
               ocr.main()
               szum_bw.main()
               szum_rgb.main()
               rozroznialnosc.main()
               
               print("###")
               print("KONIEC PRZETWARZANIA", cut(im_bef_crop.filename))
               print("###")
          
          else:
               photo_to_repair(cut(im_bef_crop.filename))
               print("NIEPOWODZENIE:", cut(im_bef_crop.filename) )
               
          # rm_files_folder("crop_pom") 
          # rm_files_folder("cropped_photo")
          # rm_files_folder("cropped_images")
          
     
    print("Koniec całego przetwarzania")
    
if __name__ == "__main__":
    main()