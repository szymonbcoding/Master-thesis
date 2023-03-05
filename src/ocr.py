import cv2 
import pytesseract
from openpyxl import load_workbook

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

def getFontSize(line: str) -> str:

    if(line[1] == ' '):
        return line[0]
    else:
        return line[:3]

def recognition(n: int) -> float:
    
    tekst = " Praca magisterska rnGC6mmmm"
    OCR1_nr_list = ["6.5", "6", "5.5", "5", "4.5", "4", "3.5", "3"]
    OCR2_nr_list = ["2.7", "2.4", "2.1", "1.8", "1.5", "1.2", "1", "0.8", "0.6"]

    img = cv2.imread("../data/cropped_images/OCR" + str(n) + ".png")

    # Adding custom options
    custom_config = r'--oem 3 --psm 6'

    f = open("../data/OCR" + str(n) +"_temp.txt", "w")
    
    try:
        if(pytesseract.image_to_string(img, config=custom_config)):
            f.write(pytesseract.image_to_string(img, config=custom_config))

            f.close()

            i=0
            label = eval("OCR" + str(n) + "_nr_list")

            with open('../data/OCR' + str(n) + '_temp.txt','r') as file: 
                for line in file:
                    if(line[0].isnumeric()):
                        value = label[i] + tekst

                        if(value in line):
                            if(i != len(label) - 1):
                                i += 1
                        else:
                            i -= 1
                            break
            if(i<0):
                return 9
            else:
                return label[i]
        else:
            return 8
    except:
        return 8
    
def main():
    
    print("Przetwaranie OCR...")
    
    wb2 = load_workbook(filename = '../data/komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 24)
    
    min = 10
    
    for i in range(1,3):
        temp = float(recognition(i))
        if(temp < min):
            min = temp
    
    if(min == 9):
        out = "Nic nie rozpoznano"
    elif(min == 8):
        out = "Blad modulu OCR"
    else:
        out = min
    
    sheet2.cell(row = 24, column = empty_col).value = out
    
    wb2.save('../data/komunikat.xlsx')
    print("OCR przetworzone")
    
    
if __name__ == "__main__":
    main()

