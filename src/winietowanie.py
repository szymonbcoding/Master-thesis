from PIL import Image
import PIL
import math

from openpyxl import load_workbook

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

def find_odd_pixel(v: int, avr1: float) -> bool:
    
    if(abs(v - avr1) > 50):
        return True
    else:
        return False

def calc_avr(photo: PIL.Image.Image) -> float:
    
    w, h = photo.size
    n = w*h

    s = 0

    for j in range(w):
        for k in range(h):
            s += photo.getpixel((j, k))

    return (s/n)

def winiet(avr_v: float, avr_c: float) -> bool:

    r = abs(avr_c-avr_v)
    
    if(r >= 10):
        return 2
    elif(10 > r > 5):
        return 1
    else:
        return 0

############################################
def main():

    print("Przetwarzanie: Winietowanie...")
    
    #komunikat dla uzytkownika
    wb2 = load_workbook(filename = '../data/komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 22)
    
    r_image_list = []
    c_image_list = []

    for i in range(1, 5):
    
        c_image_list.append(Image.open("../data/cropped_images/WINIET_C" + str(i) + ".png").convert("L"))
        r_image_list.append(Image.open("../data/cropped_images/WINIET_R" + str(i) + ".png").convert("L"))


    sc = 0
    sr = 0

    #suma wartosci pikseli centrum kadru
    for i in range(4):
        sc += calc_avr(c_image_list[i])

    #suma wartosci pikseli rogow kadru
    for i in range(4):
        sr += calc_avr(r_image_list[i])
    
    avr_vertices = sr/4

    avr_center = sc/4

    a_v = round(avr_vertices, 1)

    a_c = round(avr_center, 1)
    
    if(not winiet(avr_vertices, avr_center)):
        message = "Brak winietowania"
    elif(winiet(avr_vertices, avr_center) == 1):
        message = "Umiarkowe winietowanie"
    elif(winiet(avr_vertices, avr_center) == 2):
        message = "Widoczne winietowanie"
    else:
        message = "Błąd"
        
    sheet2.cell(row = 120, column = empty_col).value = a_c - a_v
    
    sheet2.cell(row = 22, column = empty_col).value = message
    
    wb2.save('../data/komunikat.xlsx') 
    
    print("Winietowanie: Zakonczono")
    

if __name__ == "__main__":
    #main(photo, mode)
    main()