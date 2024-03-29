from PIL import Image
import PIL
import math
from openpyxl import load_workbook

#DO ZMIANY, JEŚLI ARKUSZ WYCINANIA SIĘ ZMIENI
x_mm = 198
y_mm = 38

square_side_mm = 20

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

def find_odd_pixel(v: int, avr1: float) -> bool:
    
    if(abs(v - avr1) > 50):
        return True
    else:
        return False

def calc_avr(photo: PIL.Image.Image) -> list:
    
    avr_list = []
    w, h = photo.size
    n = w*h

    for i in range(3):
        s = 0
        for j in range(w):
            for k in range(h):
                s += photo.getpixel((j, k))[i]
        avr_list.append(s/n) 

    return avr_list 

def calc_deviation(photo: PIL.Image.Image, avr_list: list) -> list:
    
    dev_list = []

    w, h = photo.size
    n = w*h

    for i in range(3):
        s = 0
        for j in range(w):
            for k in range(h):
                s += (photo.getpixel((j, k))[i] - avr_list[i]) ** 2
        dev_list.append(math.sqrt(s/n)) 
    
    return dev_list

def calc_percent_deviation(dev_list: list, avr_list: list) -> list:

    perc_dev_list = []

    for i in range(3):
        perc_dev_list.append(dev_list[i]/avr_list[i] * 100)

    return perc_dev_list

def ca(d: list) -> float:
    s = 0
    for i in d:
        s += i
    
    return s/3
        

############################################
#def main(photo: PIL.Image.Image, mode: int) -> list:
def main():

    print("Przetwarzanie: Szum RGB...")
    
    wb = load_workbook(filename = '../data/dane_szczegolowe.xlsx')
    sheet = wb['7_Szum_RGB']
    
    empty_row = find_empty_row(sheet)
    
    #wstaw ścieżkę do zdjęcia
    im = Image.open("../data/cropped_images/CD1.png")
    x,y = im.size

    square_side_px = math.floor(x * square_side_mm / x_mm)

    px = im.convert("RGB")

    v_half = math.floor(y/2)

    #obliczanie średniej wartości pikseli pierwszego wiersza
    sum1 = 0
    
    for i in range(x):
        for j in range(3):
            sum1 += px.getpixel((i, 0))[j]

    #dzielnik razy 3, bo 3 składowe - RGB
    avr1 = math.floor(sum1/(3*x))

    h_pointer = -1

    for i in range(x):

        j = sum(px.getpixel((i, v_half)))/3

        if(find_odd_pixel(j, avr1)):
            h_pointer = i + 0.25 * square_side_px
            
            #print("avr1:", avr1, "sumoddpixel:", j)
            break
    
    crop_coords = []

    for i in range(9):

        supp_list = []

        #left, top, right, down
        supp_list.append(h_pointer + i * square_side_px)
        supp_list.append(v_half - 0.15 * square_side_px)
        supp_list.append(h_pointer + (i + 0.5) * square_side_px)
        supp_list.append(v_half + 0.15 * square_side_px)

        crop_coords.append(supp_list)

    cropped_images = []

    for i in range(9):
        cropped_images.append(px.crop((crop_coords[i][0], crop_coords[i][1], crop_coords[i][2], crop_coords[i][3])))
        cropped_images[i].save('../data/rgb/' + str(i) + '.png')

    kom_d = []
    
    for i in range(9):

        a_list = calc_avr(cropped_images[i])
        d_list = calc_deviation(cropped_images[i], a_list)
        pd_list = calc_percent_deviation(d_list, a_list)

        kom_d.append(ca(d_list))
        
        #Avr loop
        for j in range(3):
            sheet.cell(row = empty_row, column = i*9 + 1 + j * 3).value = round(a_list[j], 2)
            
        #Dev loop  
        for j in range(3):
            sheet.cell(row = empty_row, column = i*9 + 2 + j * 3).value = round(d_list[j], 2)
        
        #Pdev loop [%]
        for j in range(3):
            sheet.cell(row = empty_row, column = i*9 + 3 + j * 3).value = round(pd_list[j], 2)

    wb.save('../data/dane_szczegolowe.xlsx')
    wb.close()
    
    wb2 = load_workbook(filename = '../data/komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 26)
    
    #zapis odchyleń standardowych do arkusza komunikat
    
    for n in range(len(kom_d)):
    
        sheet2.cell(row = 26 + n, column = empty_col).value = round(kom_d[n], 2)
    
    wb2.save('../data/komunikat.xlsx')
    wb2.close()
    
    print("Szum RGB: Zakonczono")
    
if __name__ == "__main__":
    #main(photo, mode)
    main()