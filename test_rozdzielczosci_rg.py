from PIL import Image
import PIL
import math

from openpyxl import load_workbook

import glob

def openFolder(path):
    for filename in glob.glob(path + '/*.png'):
        img=Image.open(filename)

    return img

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c


def find_max_index(t: tuple) -> int:
    max = -1
    max_index = -1
    for idx, val in enumerate(t):
        if(val>max):
            max = val
            max_index = idx
            
    return max_index

def main():
    
    print("RG_RT: Przetwarzanie...")
    
    wb = load_workbook(filename = 'dane_szczegolowe.xlsx')
    sheet = wb['2b_RG_RT']
    
    empty_row = find_empty_row(sheet)
    
    #maksymalna wysokosc testu rozdzielczosci[mm]
    MAX = 20
    #wysokosc calej planszy testowej [mm]
    h = 366.67

    #zaladowanie całego obrazu planszy
    im1 = openFolder("cropped_photo")

    x_all, y_all = im1.size

    if(0.63 < y_all/x_all < 0.705):
        w = 550
    elif(0.705 <= y_all/x_all < 0.79):
        w = 488.889
    else:
        print("Błędna rozdzielczość planszy.")

    #przeliczenie 20mm (bok testu rozdzielczosci) na piksele
    px_20mm = math.floor(y_all * MAX/h)

    pict = Image.open('cropped_images/RG_RT.png')

    im2 = pict.convert("RGB")

    #rozdzielczosc wycinka
    x, y = im2.size

    #print("Green:", im2.getpixel((25, 27)))
    
    #print("Red:", im2.getpixel((42, 35)))

    #PETLA TESTU ROZDZIELCZOSCI POZIOMEJ

    #inicjalizacja zmiennych pomocniczych
    hor_res = 0
    v_pointer = -1

    #for i in range(math.floor(1.5 * px_20mm)):
    for i in range(math.floor(y / 2)):
        #resetowanie zmiennych
        hor_edges = 0
        hor_flag = False

        for j in range (0, x):

            k = im2.getpixel((j,i))
            
            if((k[1] >= 100) and find_max_index(k) == 1 and k[0] <= 150 and k[2] <= 150 and not hor_flag):

                hor_flag = True
            
            elif((k[0] >= 100) and find_max_index(k) == 0 and k[1] <= 150 and k[2] <= 150 and hor_flag):
                
                hor_flag = False
                hor_edges += 1

                if(v_pointer == -1):
                    v_pointer = i + px_20mm
            
            #sprawdzanie czy program naliczyl 5 zbocz 
            if(hor_edges>=5):
                hor_res += 1
                break

    #print("hor_res: " + str(hor_res))
    #stosunek poprawnych wierszy (najlepszy mozliwy wynik = 1, najgorszy mozliwy wynik = 0)
    p_row_found = hor_res/px_20mm

    #2 - (1.8 * 0) = 2 (najgorszy wynik)
    #2 - (1.8 * 1) = 0.2 (najlepszy wynik)
    h_mm_resolution = 2 - (1.8 * p_row_found)
    sheet.cell(row = empty_row, column = 1).value = round(h_mm_resolution, 3)
    
    real_h_px_resolution = math.floor(w/h_mm_resolution)

    #print("Rzeczywista pozioma rozdzielczosc: " + str(real_h_px_resolution))
    sheet.cell(row = empty_row, column = 2).value = real_h_px_resolution
    #print("Maksymalna mozliwa pozioma rozdzielczosc " + str(math.floor(w/0.2)))
    sheet.cell(row = empty_row, column = 3).value = math.floor(w/0.2)
    
    #PETLA TESTU ROZDZIELCZOSCI PIONOWEJ

    ver_res = 0

    for i in range(x):
        #resetowanie zmiennych
        ver_edges = 0
        ver_flag = False

        for j in range (v_pointer, y):

            k = im2.getpixel((i,j))
            
            if((k[0] >= 100) and find_max_index(k) == 0 and k[1] <= 150 and k[2] <= 150 and not ver_flag):

                ver_flag = True
            
            elif((k[1] >= 100) and find_max_index(k) == 1 and k[0] <= 150 and k[2] <= 150 and ver_flag):
                
                ver_flag = False
                ver_edges += 1
                
            #sprawdzanie czy wykryto przynajmniej 5 zbocz
            if(ver_edges>=5):
                ver_res += 1
                break

    #print("ver_res: " + str(ver_res))
    #stosunek poprawnych kolumn (najlepszy wynik = 1, najgorszy wynik = 0)
    p_col_found = ver_res/px_20mm

    #2 - (1.8 * 0) = 2 (najgorszy wynik)
    #2 - (1.8 * 1) = 0.2 (najlepszy wynik)
    v_mm_resolution = 2 - (1.8 * p_col_found)
    sheet.cell(row = empty_row, column = 4).value = round(v_mm_resolution, 3)
    real_v_px_resolution = math.floor(h/v_mm_resolution)
    sheet.cell(row = empty_row, column = 5).value = real_v_px_resolution

    #print("Rzeczywista pionowa rozdzielczosc: " + str(real_v_px_resolution))
    
    #print("Najwyzsza mozliwa pionowa rozdzielczosc " + str(math.floor(h/0.2)))
    sheet.cell(row = empty_row, column = 6).value = math.floor(h/0.2)

    wb.save('dane_szczegolowe.xlsx')
    wb.close()
    
    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 13)
    
    sheet2.cell(row = 13, column = empty_col).value = real_h_px_resolution * real_v_px_resolution
    
    sheet2.cell(row = 14, column = empty_col).value = real_h_px_resolution 
    
    sheet2.cell(row = 15, column = empty_col).value = real_v_px_resolution
    
    wb2.save('komunikat.xlsx') 
    
    print("RG_RT: Zakonczono")
    
if __name__ == "__main__":
    main()
            
