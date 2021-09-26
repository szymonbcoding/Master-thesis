import cv2 
import pytesseract
from openpyxl import load_workbook

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

def getFontSize(line: str) -> str:

    if(line[1] == ' '):
        return line[0]
    else:
        return line[:3]

def recognition(n: int) -> float:
    
    tekst = " Praca magisterska rnGC6mmmm"
    OCR1_nr_list = ["6.5", "6", "5.5", "5", "4.5", "4", "3.5", "3"]
    OCR2_nr_list = ["2.7", "2.4", "2.1", "1.8", "1.5", "1.2", "1", "0.8", "0.6"]

    #mode = "OCR1"

    img = cv2.imread("cropped_images/OCR" + str(n) + ".png")

    # Adding custom options
    custom_config = r'--oem 3 --psm 6'
    #print(pytesseract.image_to_string(img, config=custom_config))

    f = open("OCR" + str(n) +"_temp.txt", "w")
    
    try:
        if(pytesseract.image_to_string(img, config=custom_config)):
            f.write(pytesseract.image_to_string(img, config=custom_config))

            f.close()

            i=0
            label = eval("OCR" + str(n) + "_nr_list")

            with open('OCR' + str(n) + '_temp.txt','r') as file: 
                for line in file:
                    if(line[0].isnumeric()):
                        #printlabel[i] + tekst)
                        #print("line:", line)
                        value = label[i] + tekst
                        #value = "6.5 Praca magisterska rnGC6mmmm"

                        if(value in line):
                            #print(getFontSize(line) + " - ok")
                            if(i != len(label) - 1):
                                i += 1
                        else:
                            i -= 1
                            break
            if(i<0):
                return 9
            else:
                return label[i]
        else:
            return 8
    except:
        return 8
    
def main():
    
    print("Przetwaranie OCR...")
    
    wb = load_workbook(filename = 'dane_szczegolowe.xlsx')
    sheet = wb['5_OCR']
    
    empty_row = find_empty_row(sheet)
    
    min = 10
    
    for i in range(1,3):
        temp = float(recognition(i))
        if(temp < min):
            min = temp
    
    if(min == 9):
        out = "Nic nie rozpoznano"
    elif(min == 8):
        out = "Blad modulu OCR"
    else:
        out = min

    sheet.cell(row = empty_row, column = 1).value = out
    
    wb.save('dane_szczegolowe.xlsx')
    wb.close()
    
    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 24)
    
    sheet2.cell(row = 24, column = empty_col).value = out
    
    wb2.save('komunikat.xlsx')
    print("OCR przetworzone")
    
    
if __name__ == "__main__":
    main()

