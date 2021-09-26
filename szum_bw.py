from PIL import Image
import PIL
import math
from openpyxl import load_workbook

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

def calc_avr(photo: PIL.Image.Image) -> float:
    
    w, h = photo.size
    n = w*h

    s = 0
    for j in range(w):
        for k in range(h):
            s += photo.getpixel((j, k)) 

    return s/n

def calc_deviation(photo: PIL.Image.Image, avr: float) -> float:

    w, h = photo.size
    n = w*h

    s = 0
    for j in range(w):
        for k in range(h):
            s += (photo.getpixel((j, k)) - avr) ** 2
    
    return math.sqrt(s/n)

def calc_percent_deviation(dev: float, avr: float) -> float:
    return dev/avr * 100

############################################
#def main(photo: PIL.Image.Image, mode: int) -> list:
def main():
    
    wb = load_workbook(filename = 'dane_szczegolowe.xlsx')
    sheet = wb['6_Szum_BW']
    
    empty_row = find_empty_row(sheet)

    #wstaw ścieżkę do zdjęcia
    im = Image.open("cropped_images/SD.png")
    # x,y = im.size

    px = im.convert("L")

    avr_supp = calc_avr(px)
    dev_supp = calc_deviation(px, avr_supp)
    pdev_supp = calc_percent_deviation(dev_supp, avr_supp)

    d_out = round(dev_supp, 2)
    pd_out = round(pdev_supp, 2)
    
    sheet.cell(row = empty_row, column = 1).value = round(avr_supp, 2)
    sheet.cell(row = empty_row, column = 2).value = d_out
    sheet.cell(row = empty_row, column = 3).value = pd_out
    
    wb.save('dane_szczegolowe.xlsx')
    wb.close()
    
    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 25)
    
    sheet2.cell(row = 25, column = empty_col).value = d_out
    
    wb2.save('komunikat.xlsx')
    
if __name__ == "__main__":
    #main(photo, mode)
    main()

    

        






