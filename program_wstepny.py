from PIL import Image
import PIL
from openpyxl import load_workbook
import math


import kontrola_kadrowania

import glob

def openFolder(path):
    for filename in glob.glob(path + '/*.JPG'):
        img=Image.open(filename)

    return img

crop_list = []
save_list = []
labels = []

w34 = 488.889
w23 = 550
h = 366.67

x_crop_difference = 30
y_crop_difference = 51.5

def main():
    if(kontrola_kadrowania.main()):
        
        im = openFolder("photo")
        x, y = im.size

        if(0.63 < y/x < 0.705):
            mode = 23
            w = w23
        elif(0.705 <= y/x < 0.79):
            mode = 34
            w = w34
        else:
            print("Nieobslugiwana proporcja obrazu.")

        wb = load_workbook(filename = 'coordinates_v2.xlsx')

        #otworz odpowiedni arkusz
        if(mode == 23):
            sheet = wb['23']
        elif(mode == 34):
            sheet = wb['34']

        for i in range(2,19):
            crop_supp = []
            for j in range(1,6):
                if(j==1):
                    l = sheet.cell(row = i, column = 1).value
                    labels.append(l)
                else:
                    k = round(sheet.cell(row = i, column = j).value, 4)
                    if(k):
                        crop_supp.append(k)
                    if(j==5):
                        crop_list.append(crop_supp)


        for m in range(0, len(crop_list)):
            save_list.append(im.crop((math.floor(crop_list[m][0]*x/w), math.floor(crop_list[m][1]*y/h), math.floor(crop_list[m][2]*x/w), math.floor(crop_list[m][3]*y/h))))
            save_list[m].save('cropped_images/' + labels[m] + '.JPG')

    else:
        print("Blad kadrowania")
    
if __name__ == "__main__":
    main()


    

    


    
