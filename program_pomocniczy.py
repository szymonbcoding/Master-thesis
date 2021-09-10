from PIL import Image
import PIL
from openpyxl import load_workbook
import math
import glob

def openFolder(path):
    for filename in glob.glob(path + '/*.JPG'):
        img=Image.open(filename)

    return img

w34 = 488.889
w23 = 550
h = 366.67

im = openFolder("photo")
x, y = im.size

wb = load_workbook(filename = 'coordinates_v2.xlsx')

if(0.63 < y/x < 0.705):
    mode = 23
    w = w23
    sheet = wb['23']
elif(0.705 <= y/x < 0.79):
    mode = 34
    w = w34
    sheet = wb['34']
    
else:
    print("Nieobslugiwana proporcja obrazu.")

#           0           1        2         3         4        5        6        7       8      9
label = ["BW_RT1", "BW_RT2", "BW_RT3", "BW_RT4", "BW_RT5", "GB_RT", "RB_RT", "RG_RT", "CD1", "CD2"]

l = 7
d = -3
t = 1

pointer = -1

for i in range(2, 19):
    if(label[l] in sheet.cell(row = i, column = 1).value):
        pointer = i


if(pointer > 1):

    left = round(sheet.cell(row = pointer, column = 2).value, 4)
    top = round(sheet.cell(row = pointer, column = 3).value, 4)
    right = round(sheet.cell(row = pointer, column = 4).value, 4)
    down = round(sheet.cell(row = pointer, column = 5).value, 4)
    
    dright = d
    ddown = t

    im.crop(((left + d) * x/w, (top + t) * y/h, (right + dright) * x/w, (down + ddown) * y/h)).save('cropped_images/' + label[l] + '.JPG')
else:
    print("Blad")





    

    


    
