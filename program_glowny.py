import sys
#import rogi 
from PIL import Image
from PIL.ExifTags import TAGS
from openpyxl import load_workbook
import glob

#własne skrypty 
import test_rozdzielczosci_bw_mtf
import test_rozdzielczosci_gb
import test_rozdzielczosci_rb
import test_rozdzielczosci_rg

import kadruj
import program_wstepny
import winietowanie
import dystorsje
import ocr
import szum_bw
import szum_rgb
import rozroznialnosc

def openFolderMI(path: str) -> list:
    
    image_list = []
    
    for filename in glob.glob(path + '/*.JPG'):
        im = Image.open(filename)
        image_list.append(im)
        
    for filename in glob.glob(path + '/*.png'):
        im = Image.open(filename)
        image_list.append(im)

    return image_list

def openFolder(path):
    for filename in glob.glob(path + '/*.png'):
        img=Image.open(filename)

    return img

def main():
    #https://www.thepythoncode.com/article/extracting-image-metadata-in-python
    
    im_list = []
    im_list = openFolderMI("photo")
    
    for im_bef_crop in im_list:
          kadruj.main(im_bef_crop)

          image = openFolder("cropped_photo")
          
          x, y = image.size
               
          # extract EXIF data
          exifdata = image.getexif()

          #data = [exifdata[0x0110], exifdata[0xA405], exifdata[0x0100], 
          # exifdata[0x0101], exifdata[0x9202][0]/exifdata[0x9202][1], 
          # exifdata[0x9201][0]/exifdata[0x9201][1], exifdata[0x8827], "-" ]

          wb = load_workbook(filename = 'output.xlsx')

          sheet = wb['1_Ustawienia_aparatu']

          empty_row = 0

          for i in range(4, 1000):
               if(not (sheet.cell(row = i, column = 1).value)):
                    empty_row = i
                    break
               
          #Model aparatu
          try:
               model = exifdata[0x0110]
               sheet.cell(row = empty_row, column = 1).value =  model
          except:
               model = "-"
               sheet.cell(row = empty_row, column = 1).value = model
               
          #Dlugosc ogniskowa
          try:
               sheet.cell(row = empty_row, column = 2).value = exifdata[0xA405] 
          except:
               sheet.cell(row = empty_row, column = 2).value = "-"
               
          #Długość obrazu
          try:
               sheet.cell(row = empty_row, column = 3).value = exifdata[0x0100] 
          except:
               sheet.cell(row = empty_row, column = 3).value = x
               
          #Szerokość obrazu
          try:
               sheet.cell(row = empty_row, column = 4).value = exifdata[0x0101] 
          except:
               sheet.cell(row = empty_row, column = 4).value = y
               
          #Przyslona
          try:
               sheet.cell(row = empty_row, column = 5).value = exifdata[0x9202][0]/exifdata[0x9202][1] 
          except:
               sheet.cell(row = empty_row, column = 5).value = "5.6"
               
          #Czas naswietlania
          try:
               sheet.cell(row = empty_row, column = 6).value = exifdata[0x9201][0]/exifdata[0x9201][1] 
          except:
               sheet.cell(row = empty_row, column = 6).value = "1/100"
               
          #ISO
          try:
               sheet.cell(row = empty_row, column = 7).value = exifdata[0x8827] 
          except:
               sheet.cell(row = empty_row, column = 7).value = "200"
               
          #Model obiektywu
          try:
               sheet.cell(row = empty_row, column = 8).value = exifdata[0xA434] 
          except:
               sheet.cell(row = empty_row, column = 8).value = "-"
               
          #Nazwa zdjecia
          try:
               sheet.cell(row = empty_row, column = 9).value = image.filename
          except:
               sheet.cell(row = empty_row, column = 9).value = "-"

                         
          wb.save('output.xlsx')
          wb.close()

          wb2 = load_workbook(filename = 'komunikat.xlsx')

          sheet2 = wb2['Arkusz1']

          empty_col = 0

          for i in range(2, 1000):
               if(not (sheet.cell(row = 1, column = i).value)):
                    empty_col = i
                    break
          #model
          sheet2.cell(row = 1, column = empty_col).value = model

          #Nazwa zdjecia
          sheet2.cell(row = 2, column = empty_col).value = image.filename
          
          #rozdzielczosc deklarowana przez producenta
          sheet2.cell(row = 3, column = empty_col).value = x * y

          wb2.save('komunikat.xlsx')

          program_wstepny.main()
          test_rozdzielczosci_bw_mtf.main()
          test_rozdzielczosci_rg.main()
          test_rozdzielczosci_gb.main()
          test_rozdzielczosci_rb.main()
          winietowanie.main()
          dystorsje.main() #do dopracowania
          ocr.main()
          szum_bw.main()
          szum_rgb.main()
          rozroznialnosc.main()
          print("Koniec przetwarzania", image.filename)
     
    print("Koniec całego przetwarzania")
    
if __name__ == "__main__":
    main()