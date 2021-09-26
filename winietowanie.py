from PIL import Image
import PIL
import math

from openpyxl import load_workbook

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

def find_odd_pixel(v: int, avr1: float) -> bool:
    
    if(abs(v - avr1) > 50):
        return True
    else:
        return False

def calc_avr(photo: PIL.Image.Image) -> float:
    
    w, h = photo.size
    n = w*h

    s = 0

    for j in range(w):
        for k in range(h):
            s += photo.getpixel((j, k))

    return (s/n)

def winiet(avr_v: float, avr_c: float) -> bool:

    r = abs(avr_c-avr_v)
    
    if(r >= 10):
        return 2
    elif(10 > r > 5):
        return 1
    else:
        return 0

############################################
def main():

    print("Przetwarzanie: Winietowanie...")
    
    wb = load_workbook(filename = 'dane_szczegolowe.xlsx')
    sheet = wb['3_Winietowanie']
    
    empty_row = find_empty_row(sheet)
    
    image_list = []

    for i in range(1, 6):
    
        image_list.append(Image.open("cropped_images/BW_RT" + str(i) + ".png").convert("L"))

    sv = 0

    for i in range(4):
        sv += calc_avr(image_list[i])
    
    avr_vertexes = sv/4

    avr_center = calc_avr(image_list[4])

    print("Srednia wartosc piksela na wierzcholkach planszy:", avr_vertexes)
    a_v = round(avr_vertexes, 1)
    sheet.cell(row = empty_row, column = 1).value = a_v
    
    print("Srednia wartosc piksela na srodku planszy:", avr_center)
    a_c = round(avr_center, 1)
    sheet.cell(row = empty_row, column = 2).value = a_c
    
    if(not winiet(avr_vertexes, avr_center)):
        message = "Brak winietowania"
    elif(winiet(avr_vertexes, avr_center) == 1):
        message = "Umiarkowe winietowanie"
    elif(winiet(avr_vertexes, avr_center) == 2):
        message = "Widoczne winietowanie"
    else:
        message = "Błąd"
        
    sheet.cell(row = empty_row, column = 3).value = message
    
    wb.save('dane_szczegolowe.xlsx')
    wb.close()
    
    #komunikat dla uzytkownika
    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 22)
    
    #sheet2.cell(row = 9, column = empty_col).value = abs(a_c - a_v)
    
    sheet2.cell(row = 22, column = empty_col).value = message
    
    wb2.save('komunikat.xlsx') 
    
    print("Winietowanie: Zakonczono")
    

if __name__ == "__main__":
    #main(photo, mode)
    main()