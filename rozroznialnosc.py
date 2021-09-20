from PIL import Image
import PIL
import math
from openpyxl import load_workbook

#DO ZMIANY, JEŚLI ARKUSZ SIĘ ZMIENI
x_mm = 118
y_mm = 58

def find_empty_row(sh) -> int:
    
    r = 0
    
    for i in range(4, 1000):
        if(not (sh.cell(row = i, column = 1).value)):
            r = i
            break
    return r

def find_empty_col(sh, p: int) -> int:
    
    c = 0
    
    for i in range(2, 1000):
        if(not (sh.cell(row = p, column = i).value)):
            c = i
            break
    
    return c

square_side_mm = 20

def find_odd_pixel(v: int, avr1: float) -> bool:
    
    if(abs(v - avr1) > 50):
        return True
    else:
        return False

def calc_avr(photo: PIL.Image.Image) -> float:
    
    w, h = photo.size
    n = w*h

    s = 0

    for j in range(w):
        for k in range(h):
            s += photo.getpixel((j, k))

    return (s/n)

############################################
#def main(photo: PIL.Image.Image, mode: int) -> list:
def main():

    print("Przetwarzanie: Rozroznialnosc...")
    
    wb = load_workbook(filename = 'output.xlsx')
    sheet = wb['8_Rozroznialnosc']
    
    empty_row = find_empty_row(sheet)
    
    #wstaw ścieżkę do zdjęcia
    im = Image.open("cropped_images/CD2.png")
    x,y = im.size

    square_side_px = math.floor(x * square_side_mm / x_mm)

    px = im.convert("L")

    v_quarter = math.floor(y/4)

    #obliczanie średniej wartości pikseli pierwszego wiersza
    sum1 = 0
    for i in range(x):
            sum1 += px.getpixel((i, 0))

    avr1 = sum1/x

    h_pointer = -1

    for i in range(x):

        j = px.getpixel((i, v_quarter))

        if(find_odd_pixel(j, avr1)):
            h_pointer = i + 0.25 * square_side_px
            break
    
    crop_coords = []

    for i in range(10):

        supp_list = []

        if(i < 5):
            q = 1
        elif(i >= 5):
            q = 3

        #left, top, right, down
        supp_list.append(h_pointer + (i % 5) * square_side_px)
        supp_list.append(q * v_quarter - 0.15 * square_side_px)
        supp_list.append(h_pointer + ((i % 5) + 0.5) * square_side_px)
        supp_list.append(q * v_quarter + 0.15 * square_side_px)

        crop_coords.append(supp_list)
        
    cropped_images = []

    for i in range(10):
        cropped_images.append(px.crop((crop_coords[i][0], crop_coords[i][1], crop_coords[i][2], crop_coords[i][3])))
        #cropped_images[i].save('bw/' + str(i) + '.JPG')
   
    good = 4
    notbad = 1
    
    werdykt_b = -1

    #czarne
    for i in range(5):

        a = calc_avr(cropped_images[i])
        
        if(i>0):
            ab_r = abs(a - calc_avr(cropped_images[i - 1]))
            if(ab_r > good):
                werdykt_b += 3
            elif(ab_r > notbad):
                werdykt_b += 1
            sheet.cell(row = empty_row, column = i).value = round(ab_r, 2)
            
        #sheet.cell(row = empty_row, column = i + 1).value = round(a, 2)
    
    if(werdykt_b >= 11):
        m_b = "Zadowalająca rozróżnialność"
    elif(werdykt_b >= 3):
        m_b = "Dopuszczająca rozróżnialność"
    else:
        m_b = "Niedostateczna rozróżnialność"
        
    sheet.cell(row = empty_row, column = 5).value = m_b
    
    werdykt_w = -1
    
    #biale
    for i in range(5):
        
        a = calc_avr(cropped_images[i + 5])

        if(i>0):
            aw_r = abs(a - calc_avr(cropped_images[i + 4]))
            if(aw_r > good):
                werdykt_w += 3
            elif(aw_r > notbad):
                werdykt_w += 1
            sheet.cell(row = empty_row, column = i + 5).value = round(aw_r, 2)
        
    
    if(werdykt_w >= 11):
        m_w = "Zadowalająca rozróżnialność"
    elif(werdykt_w >= 3):
        m_w = "Dopuszczająca rozróżnialność"
    else:
        m_w = "Niedostateczna rozróżnialność"
    sheet.cell(row = empty_row, column = 10).value = m_w
    
    wb.save('output.xlsx')
    wb.close()
    wb2 = load_workbook(filename = 'komunikat.xlsx')
    
    sheet2 = wb2['Arkusz1']
    
    empty_col = find_empty_col(sheet2, 35)
    
    sheet2.cell(row = 35, column = empty_col).value = m_b
    
    sheet2.cell(row = 36, column = empty_col).value = m_w
    
    wb2.save('komunikat.xlsx')
    
    print("Rozroznialnosc przetworzona")
    

if __name__ == "__main__":
    main()